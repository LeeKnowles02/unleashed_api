from dotenv import load_dotenv
from db import start_run, finish_run

load_dotenv()

from flask import Flask, render_template, request, Response
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import io
from typing import Any, Dict, Optional, List

from config import Config
from unleashed_client import UnleashedClient
from exports import (
    sales_orders,
    customers,
    suppliers,
    products,
    warehouses,
    stock_on_hand,
    sales_shipments,
    credit_notes,
    invoices,
)

cfg = Config()

client = UnleashedClient(
    base_url=cfg.UNLEASHED_BASE_URL,
    api_id=cfg.UNLEASHED_API_ID,
    api_key=cfg.UNLEASHED_API_KEY,
    client_type=cfg.UNLEASHED_CLIENT_TYPE,
    timeout_seconds=cfg.REQUEST_TIMEOUT_SECONDS,
)

EXPORTS = {
    "products_api": {
        "category": "products",
        "label": "Products",
        "description": "Product master data",
        "sheet_name": "Products",
        "dummy": products.dummy,
        "api": lambda: products.from_api(client),
    },
    "invoices": {
        "category": "sales",
        "label": "Invoices",
        "description": "Revenue documents (header-only for now)",
        "sheet_name": "Invoices",
        "dummy": invoices.dummy,
        "api": lambda: invoices.from_api(client),
    },
    "credit_notes": {
        "category": "sales",
        "label": "Credit Notes",
        "description": "Returns and revenue corrections",
        "sheet_name": "CreditNotes",
        "dummy": credit_notes.dummy,
        "api": lambda: credit_notes.from_api(client),
    },
    "warehouses": {
        "category": "inventory",
        "label": "Warehouses",
        "description": "Warehouse master data",
        "sheet_name": "Warehouses",
        "api": lambda: warehouses.from_api(client),
    },
    "sales_shipments": {
        "category": "sales",
        "label": "Sales Shipments",
        "description": "Dispatch / fulfilment documents",
        "sheet_name": "SalesShipments",
        "dummy": sales_shipments.dummy,
        "api": lambda: sales_shipments.from_api(client),
    },
    "stock_on_hand_api": {
        "category": "inventory",
        "label": "Stock On Hand (API)",
        "description": "Inventory snapshot (by product/warehouse)",
        "sheet_name": "StockOnHand",
        "api": lambda: stock_on_hand.from_api(client),
    },
    "sales_orders": {
        "category": "sales",
        "label": "Sales Orders",
        "description": "Transactional",
        "sheet_name": "SalesOrders",
        "dummy": sales_orders.dummy,
        "api": lambda **kwargs: sales_orders.from_api(client, **kwargs),
    },
    "customers": {
        "category": "customers",
        "label": "Customers",
        "description": "Customer master data",
        "sheet_name": "Customers",
        "dummy": customers.dummy,
        "api": lambda: customers.from_api(client),
    },
    "suppliers": {
        "category": "purchasing",
        "label": "Suppliers",
        "description": "Supplier master data",
        "sheet_name": "Suppliers",
        "dummy": suppliers.dummy,
        "api": lambda: suppliers.from_api(client),
    },
}


def build_workbook(selected_keys):
    wb = Workbook()
    wb.remove(wb.active)

    run_id = None
    if cfg.USE_UNLEASHED_API and client.is_configured():
        try:
            run_id = start_run(company_id="unleashed_client_1")
        except Exception:
            run_id = None

    try:
        for key in selected_keys:
            if key not in EXPORTS:
                continue

            export = EXPORTS.get(key)

            if not export:
                continue

            if "generator" in export:
                sheet_name, headers, rows = export["generator"]()
            else:
                if not cfg.USE_UNLEASHED_API or not client.is_configured():
                    sheet_name, headers, rows = export["dummy"]()
                else:
                    api_fn = export["api"]

                    try:
                        sheet_name, headers, rows = api_fn(
                            run_id=run_id, company_id="unleashed_client_1"
                        )
                    except TypeError:
                        sheet_name, headers, rows = api_fn()

            ws = wb.create_sheet(title=sheet_name[:31])
            ws.append(headers)
            for r in rows:
                ws.append(r)

            for cell in ws[1]:
                cell.font = cell.font.copy(bold=True)

            for col_idx in range(1, len(headers) + 1):
                col_letter = get_column_letter(col_idx)
                max_len = 0
                for cell in ws[col_letter]:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

        if run_id:
            finish_run(run_id, "SUCCESS")

        return wb

    except Exception as e:
        if run_id:
            finish_run(run_id, "FAILED", notes=str(e))
        raise


def run_export(key: str, **kwargs):
    export = EXPORTS.get(key)
    if not export:
        raise KeyError(key)

    if "generator" in export:
        return export["generator"]()

    if not cfg.USE_UNLEASHED_API or not client.is_configured():
        return export["dummy"]()

    return export["api"](**kwargs)


app = Flask(__name__, template_folder="templates", static_folder="static")


def build_exports_list(category: Optional[str] = None):
    exports_list = []
    for key, meta in EXPORTS.items():
        if category and meta.get("category") != category:
            continue

        exports_list.append(
            {
                "key": key,
                "label": meta.get("label", key),
                "description": meta.get("description", ""),
                "last_run": "—",
                "status": "Idle",
                "status_class": "pill-neutral",
            }
        )
    return exports_list


@app.route("/")
def index():
    return render_template(
        "index.html",
        active_page="dashboard",
        page_title="Dashboard",
        page_subtitle="Run exports and download Excel-ready files.",
        exports=build_exports_list(),
    )


@app.route("/exports/<category>")
def exports_by_category(category: str):
    titles = {
        "sales": ("Sales & Revenue", "Sales documents and revenue drivers."),
        "customers": (
            "Customers & Sales Dimensions",
            "Customer master data and segmentation.",
        ),
        "products": (
            "Products & Cost Foundation",
            "Product master, groups, and pricing.",
        ),
        "inventory": ("Inventory", "Stock position and inventory movements."),
        "purchasing": ("Purchasing", "Supplier master and purchasing documents."),
    }
    title, subtitle = titles.get(
        category, ("Exports", "Run exports and download Excel-ready files.")
    )

    return render_template(
        "index.html",
        active_page=category,
        page_title=title,
        page_subtitle=subtitle,
        exports=build_exports_list(category=category),
    )


@app.route("/ui")
def ui_showcase():
    return render_template("ui_showcase.html", active_page="ui")


@app.route("/api-status")
def api_status():
    return {
        "use_unleashed_api": cfg.USE_UNLEASHED_API,
        "configured": client.is_configured(),
        "base_url": client.base_url,
        "client_type": client.client_type,
    }


@app.route("/run-selected", methods=["POST"])
def run_selected():
    selected = request.form.getlist("exports")
    if not selected:
        return "No exports selected", 400

    wb = build_workbook(selected)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=unleashed_exports.xlsx"},
    )


@app.route("/run-single", methods=["POST"])
def run_single():
    key = request.form.get("export")
    if key not in EXPORTS:
        return "Invalid export", 400

    wb = build_workbook([key])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={key}.xlsx"},
    )


if __name__ == "__main__":
    app.run(debug=True, use_reloader=False)
